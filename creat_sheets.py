from openpyxl import load_workbook

book = load_workbook('StatATS32_0221.xlsx')
for x in range(1,29):
    book.create_sheet(str('{:02d}'.format(x)) + '.02.2021')

book.save('StatATS32_0221.xlsx')
book.close()

